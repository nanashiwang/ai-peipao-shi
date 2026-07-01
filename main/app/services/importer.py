"""导入服务。

负责把 CSV / XLSX 的原始行数据转成数据库里的家庭和消息记录。
"""

import csv
from datetime import datetime
from io import BytesIO, StringIO
from pathlib import Path
import re

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models import Family, RawMessage
from app.services.scenario import detect_checkin


FIELD_ALIASES = {
    "family_id": ["family_id", "家庭编号", "家庭ID"],
    "parent_nickname": ["parent_nickname", "家长昵称"],
    "child_grade": ["child_grade", "孩子年级"],
    "course_stage": ["course_stage", "课程阶段"],
    "unit_progress": ["unit_progress", "Unit进度", "Unit 进度", "单元进度"],
    "pbl_count": ["pbl_count", "PBL次数", "PBL 次数"],
    "checkin_rate": ["checkin_rate", "打卡率", "打卡完成率"],
    "next_milestone": ["next_milestone", "下一里程碑", "下个里程碑"],
    "campus_name": ["campus_name", "校区", "所属校区", "机构校区"],
    "coach_name": ["coach_name", "陪跑师"],
    "service_status": ["service_status", "服务状态"],
    "message_time": ["message_time", "聊天时间", "时间"],
    "speaker": ["speaker", "说话人"],
    "content": ["content", "消息内容", "内容"],
    "source": ["source", "群/单聊来源", "来源"],
    "checkin_status": ["checkin_status", "打卡状态"],
    "parent_phone": ["parent_phone", "手机号", "手机", "家长手机号"],
}

IMPORT_TEMPLATES = {
    "student_info_v1": {
        "key": "student_info_v1",
        "template_family": "student_info",
        "version": "1.0",
        "name": "学员信息模板",
        "business_type": "学员信息",
        "description": "用于初始化家庭、学员、陪跑师、课程阶段等基础档案。",
        "required_fields": ["family_id", "parent_nickname", "child_grade", "coach_name"],
        "optional_fields": ["parent_phone", "child_name", "campus_name", "course_stage", "service_status"],
        "headers": ["family_id", "parent_nickname", "parent_phone", "child_name", "child_grade", "campus_name", "coach_name", "course_stage", "service_status"],
        "sample_rows": [],
    },
    "chat_messages_v1": {
        "key": "chat_messages_v1",
        "template_family": "chat_messages",
        "version": "1.0",
        "name": "聊天记录模板",
        "business_type": "聊天记录",
        "description": "用于导入企业微信群聊/单聊消息，是画像、周报、回复和待办判断的主数据源。",
        "required_fields": ["family_id", "message_time", "speaker", "content"],
        "optional_fields": ["parent_nickname", "child_grade", "campus_name", "coach_name", "source", "checkin_status"],
        "headers": ["family_id", "parent_nickname", "child_grade", "campus_name", "coach_name", "message_time", "speaker", "content", "source", "checkin_status"],
        "sample_rows": [],
    },
    "checkin_records_v1": {
        "key": "checkin_records_v1",
        "template_family": "checkin_records",
        "version": "1.0",
        "name": "打卡记录模板",
        "business_type": "打卡记录",
        "description": "用于单独导入打卡、PBL提交或任务完成证据。",
        "required_fields": ["family_id", "message_time", "checkin_status", "content"],
        "optional_fields": ["parent_nickname", "source", "campus_name", "coach_name"],
        "headers": ["family_id", "parent_nickname", "message_time", "checkin_status", "content", "source", "campus_name", "coach_name"],
        "sample_rows": [],
    },
    "leave_makeup_v1": {
        "key": "leave_makeup_v1",
        "template_family": "leave_makeup",
        "version": "1.0",
        "name": "请假补课模板",
        "business_type": "请假缺课记录",
        "description": "用于登记请假、缺课、补课计划和跟进责任人。",
        "required_fields": ["family_id", "parent_nickname", "leave_time", "leave_reason", "makeup_plan"],
        "optional_fields": ["owner", "status", "campus_name", "coach_name"],
        "headers": ["family_id", "parent_nickname", "leave_time", "leave_reason", "makeup_plan", "owner", "status", "campus_name", "coach_name"],
        "sample_rows": [],
    },
    "course_stage_v1": {
        "key": "course_stage_v1",
        "template_family": "course_stage",
        "version": "1.0",
        "name": "课程阶段模板",
        "business_type": "课程阶段数据",
        "description": "用于导入课程阶段、Unit进度、PBL次数、打卡率和下一里程碑。",
        "required_fields": ["family_id", "parent_nickname", "course_stage", "unit_progress"],
        "optional_fields": ["child_grade", "pbl_count", "checkin_rate", "next_milestone", "campus_name", "coach_name"],
        "headers": ["family_id", "parent_nickname", "child_grade", "campus_name", "course_stage", "unit_progress", "pbl_count", "checkin_rate", "next_milestone", "coach_name"],
        "sample_rows": [],
    },
}

MAX_IMPORT_ISSUES = 100
VALID_MOBILE_RE = re.compile(r"^1[3-9]\d{9}$")


def list_import_templates() -> list[dict]:
    return [IMPORT_TEMPLATES[key] for key in sorted(IMPORT_TEMPLATES)]


def get_import_template(template_key: str) -> dict:
    template = IMPORT_TEMPLATES.get((template_key or "").strip())
    if not template:
        raise KeyError(template_key)
    return template


def import_template_csv_bytes(template_key: str) -> bytes:
    template = get_import_template(template_key)
    buf = StringIO()
    writer = csv.DictWriter(buf, fieldnames=template["headers"], lineterminator="\n")
    writer.writeheader()
    for row in template["sample_rows"]:
        writer.writerow(row)
    return buf.getvalue().encode("utf-8-sig")


# 按字段别名读取值，统一处理中文表头和英文表头。
def _get(row: dict, field: str, default: str = "") -> str:
    for key in FIELD_ALIASES[field]:
        if key in row and row[key] is not None:
            return str(row[key]).strip()
    return default


# 把导入文本转成 datetime；没有值时回退到当前时间。
def _parse_time(value: str) -> datetime:
    parsed, _ = _parse_time_with_issue(value)
    return parsed


def _parse_time_with_issue(value: str) -> tuple[datetime, str]:
    if not value:
        return datetime.utcnow(), "missing"
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y/%m/%d %H:%M", "%Y/%m/%d"):
        try:
            return datetime.strptime(value, fmt), ""
        except ValueError:
            pass
    return datetime.utcnow(), "invalid"


def _issue(row_number: int, field: str, code: str, message: str, severity: str = "error") -> dict:
    return {
        "row": row_number,
        "field": field,
        "code": code,
        "severity": severity,
        "message": message,
    }


def _looks_mojibake(text: str) -> bool:
    if "\ufffd" in text:
        return True
    question_count = text.count("?")
    if re.search(r"\?{4,}", text) or (question_count >= 6 and question_count / max(len(text), 1) >= 0.2):
        return True
    mojibake_hits = sum(1 for token in ("锛", "涓", "浠", "寰", "绯", "璇", "鎴", "鐨", "瀹") if token in text)
    return mojibake_hits >= 3


def _parse_optional_int(value: str) -> tuple[int | None, str]:
    if not value:
        return None, ""
    text = str(value).strip()
    if not text:
        return None, ""
    match = re.search(r"\d+", text)
    if not match:
        return None, "invalid"
    return int(match.group(0)), ""


def _apply_family_fields(family: Family, data: dict) -> None:
    for field in ("parent_nickname", "child_grade", "campus_name", "coach_name"):
        value = data.get(field, "")
        if value and not getattr(family, field):
            setattr(family, field, value)

    for field in ("course_stage", "unit_progress", "checkin_rate", "next_milestone", "service_status"):
        value = data.get(field, "")
        if value:
            setattr(family, field, value)

    if data.get("pbl_count") is not None:
        family.pbl_count = data["pbl_count"]


def validate_import_row(row: dict, row_number: int) -> tuple[dict, list[dict]]:
    family_id = _get(row, "family_id")
    parent_nickname = _get(row, "parent_nickname")
    content = _get(row, "content")
    speaker = _get(row, "speaker")
    source_raw = _get(row, "source")
    source = source_raw or "导入"
    parent_phone = _get(row, "parent_phone")
    message_time_raw = _get(row, "message_time")
    checkin_status = _get(row, "checkin_status")
    has_message = bool(content)
    has_message_markers = has_message or bool(message_time_raw) or bool(speaker) or bool(source_raw) or bool(checkin_status)
    parsed_time, time_issue = _parse_time_with_issue(message_time_raw) if has_message_markers else (datetime.utcnow(), "")
    pbl_count, pbl_issue = _parse_optional_int(_get(row, "pbl_count"))
    issues: list[dict] = []

    if not family_id:
        issues.append(_issue(row_number, "family_id", "missing_family_id", "家庭编号不能为空"))
    elif len(family_id) > 64 or any(ch in family_id for ch in "\r\n\t"):
        issues.append(_issue(row_number, "family_id", "invalid_family_id", "家庭编号过长或包含控制字符"))

    if not parent_nickname:
        issues.append(_issue(row_number, "parent_nickname", "missing_conversation", "家长昵称/企微会话名为空，已用家庭编号兜底", "warning"))
        parent_nickname = family_id

    if not content and has_message_markers:
        issues.append(_issue(row_number, "content", "empty_content", "消息内容不能为空"))
    elif content and _looks_mojibake(content):
        issues.append(_issue(row_number, "content", "mojibake_content", "消息内容疑似乱码"))

    if has_message and not speaker:
        issues.append(_issue(row_number, "speaker", "missing_speaker", "说话人为空，会影响上下文判断", "warning"))

    if time_issue == "missing":
        issues.append(_issue(row_number, "message_time", "missing_time", "聊天时间为空，已使用导入时间", "warning"))
    elif time_issue == "invalid":
        issues.append(_issue(row_number, "message_time", "invalid_time", "聊天时间格式无法识别"))

    if parent_phone and not VALID_MOBILE_RE.fullmatch(parent_phone):
        issues.append(_issue(row_number, "parent_phone", "invalid_phone", "手机号格式不合法", "warning"))

    if pbl_issue:
        issues.append(_issue(row_number, "pbl_count", "invalid_pbl_count", "PBL次数无法识别，已跳过该字段", "warning"))

    return {
        "family_id": family_id,
        "parent_nickname": parent_nickname,
        "child_grade": _get(row, "child_grade"),
        "course_stage": _get(row, "course_stage"),
        "unit_progress": _get(row, "unit_progress"),
        "pbl_count": pbl_count if not pbl_issue else None,
        "checkin_rate": _get(row, "checkin_rate"),
        "next_milestone": _get(row, "next_milestone"),
        "campus_name": _get(row, "campus_name"),
        "coach_name": _get(row, "coach_name"),
        "service_status": _get(row, "service_status"),
        "has_message": has_message,
        "message_time": parsed_time,
        "speaker": speaker,
        "content": content,
        "source": source,
        "checkin_status": checkin_status or detect_checkin(content),
    }, issues


# 支持 CSV 和 XLSX 两种上传格式。
def rows_from_upload(filename: str, data: bytes) -> list[dict]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        text = data.decode("utf-8-sig")
        return list(csv.DictReader(StringIO(text)))
    if suffix in {".xlsx", ".xlsm"}:
        wb = load_workbook(BytesIO(data), data_only=True)
        ws = wb.active
        headers = [str(cell.value).strip() if cell.value is not None else "" for cell in next(ws.iter_rows(max_row=1))]
        rows = []
        for raw in ws.iter_rows(min_row=2, values_only=True):
            rows.append({headers[i]: raw[i] for i in range(len(headers))})
        return rows
    raise ValueError("只支持 CSV / XLSX")


# 把导入行写入数据库，同时自动补齐家庭信息和打卡状态。
def import_rows(db: Session, rows: list[dict]) -> dict:
    families_seen = set()
    family_cache = {family.family_id: family for family in db.query(Family).all()}
    seen_keys = set()
    issues: list[dict] = []
    imported = 0
    profile_rows = 0
    skipped = 0
    for index, row in enumerate(rows, start=2):
        data, row_issues = validate_import_row(row, index)
        blocking = [item for item in row_issues if item["severity"] == "error"]
        if blocking:
            skipped += 1
            issues.extend(row_issues)
            continue
        family_id = data["family_id"]
        content = data["content"]
        if data["has_message"]:
            duplicate_key = (family_id, data["message_time"].isoformat(), data["speaker"], content)
            if duplicate_key in seen_keys:
                skipped += 1
                issues.append(_issue(index, "content", "duplicate_in_file", "同一导入文件中存在重复消息", "warning"))
                continue
            seen_keys.add(duplicate_key)
            exists = (
                db.query(RawMessage)
                .filter(
                    RawMessage.family_id == family_id,
                    RawMessage.message_time == data["message_time"],
                    RawMessage.speaker == data["speaker"],
                    RawMessage.content == content,
                )
                .first()
            )
            if exists:
                skipped += 1
                issues.append(_issue(index, "content", "duplicate_existing", "数据库中已存在相同消息，已跳过", "warning"))
                continue
        family = family_cache.get(family_id)
        if not family:
            family = Family(
                family_id=family_id,
                parent_nickname=data["parent_nickname"],
                child_grade=data["child_grade"],
                course_stage=data["course_stage"],
                unit_progress=data["unit_progress"],
                pbl_count=data["pbl_count"] if data["pbl_count"] is not None else 0,
                checkin_rate=data["checkin_rate"],
                next_milestone=data["next_milestone"],
                campus_name=data["campus_name"],
                coach_name=data["coach_name"],
                service_status=data["service_status"] or "试点中",
            )
            db.add(family)
            family_cache[family_id] = family
        else:
            _apply_family_fields(family, data)

        if data["has_message"]:
            db.add(
                RawMessage(
                    family_id=family_id,
                    message_time=data["message_time"],
                    speaker=data["speaker"],
                    content=content,
                    source=data["source"],
                    checkin_status=data["checkin_status"],
                    is_effective="Y" if len(content) >= 2 else "N",
                )
            )
            imported += 1
        else:
            profile_rows += 1
        issues.extend(row_issues)
        families_seen.add(family_id)
    db.commit()
    return {
        "families": len(families_seen),
        "messages": imported,
        "profile_rows": profile_rows,
        "skipped": skipped,
        "issue_count": len(issues),
        "issues": issues[:MAX_IMPORT_ISSUES],
    }
